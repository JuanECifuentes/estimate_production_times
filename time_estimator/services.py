"""
Business logic for time estimation processing.
This module handles Excel parsing, validation, and calculation of production metrics.
"""

import pandas as pd
import io
import math
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings


class ExcelProcessingError(Exception):
    """Custom exception for Excel processing errors"""
    pass


class TimeEstimationService:
    """
    Service class for processing time-study Excel files and generating production estimates.
    """
    
    REQUIRED_COLUMNS = ['Operador', 'Operación', 'Máquina', 'Suplemento']
    TIME_COLUMN_PREFIX = 'Tiempo'
    
    def __init__(self, working_hours_per_day: int = None):
        config = settings.TIME_ESTIMATION_CONFIG
        self.working_hours_per_day = working_hours_per_day or config.get('WORKING_HOURS_PER_DAY', 8)
    
    def validate_file_extension(self, filename: str) -> None:
        """Validate that the file has an allowed extension"""
        if not filename.lower().endswith('.xlsx'):
            raise ExcelProcessingError(
                "Invalid file format. Only .xlsx files are allowed."
            )
    
    def validate_file_size(self, file_size: int) -> None:
        """Validate that the file size is within limits"""
        max_size = settings.TIME_ESTIMATION_CONFIG.get('MAX_FILE_SIZE_MB', 10) * 1024 * 1024
        if file_size > max_size:
            raise ExcelProcessingError(
                f"File size exceeds maximum allowed size of {max_size / (1024 * 1024)}MB"
            )
    
    def read_excel_file(self, file_obj) -> pd.DataFrame:
        """
        Read Excel file and return DataFrame
        """
        try:
            df = pd.read_excel(file_obj, engine='openpyxl')
            return df
        except Exception as e:
            raise ExcelProcessingError(f"Failed to read Excel file: {str(e)}")
    
    def validate_dataframe_structure(self, df: pd.DataFrame) -> None:
        """
        Validate that the DataFrame has the required columns
        """
        columns = df.columns.tolist()
        
        # Check for required columns
        missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in columns]
        if missing_columns:
            raise ExcelProcessingError(
                f"Missing required columns: {', '.join(missing_columns)}. "
                f"Required columns: {', '.join(self.REQUIRED_COLUMNS)}"
            )
        
        # Check for at least one time column
        time_columns = [col for col in columns if self.TIME_COLUMN_PREFIX in str(col)]
        if not time_columns:
            raise ExcelProcessingError(
                f"No time measurement columns found. Expected columns starting with '{self.TIME_COLUMN_PREFIX}'"
            )
    
    def extract_time_columns(self, df: pd.DataFrame) -> List[str]:
        """
        Extract column names that contain time measurements
        """
        return [col for col in df.columns if self.TIME_COLUMN_PREFIX in str(col)]
    
    def calculate_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate production metrics for each operation.
        
        Returns a DataFrame with calculated metrics per operation.
        """
        time_columns = self.extract_time_columns(df)
        
        if not time_columns:
            raise ExcelProcessingError("No time measurement columns found")
        
        results = []
        
        for idx, row in df.iterrows():
            try:
                operator = row['Operador']
                operation = row['Operación']
                machine = row['Máquina']
                supplement_pct = float(row['Suplemento']) if pd.notna(row['Suplemento']) else 0.0
                
                # Extract time measurements
                time_values = []
                for col in time_columns:
                    value = row[col]
                    if pd.notna(value):
                        try:
                            time_values.append(float(value))
                        except (ValueError, TypeError):
                            continue
                
                if not time_values:
                    raise ExcelProcessingError(
                        f"No valid time measurements for operation '{operation}' (row {idx + 2})"
                    )
                
                # Calculate average observed time
                avg_time_seconds = sum(time_values) / len(time_values)
                
                # Calculate standard time with supplement
                supplement_multiplier = 1 + (supplement_pct / 100)
                standard_time_seconds = avg_time_seconds * supplement_multiplier
                standard_time_minutes = standard_time_seconds / 60
                
                # Calculate target units per hour and per day (floor rounding)
                units_per_hour = 3600 / standard_time_seconds if standard_time_seconds > 0 else 0
                units_per_day = units_per_hour * self.working_hours_per_day
                
                results.append({
                    'Operación': operation,
                    'Máquina': machine,
                    'Tiempo Promedio (seg)': round(avg_time_seconds, 2),
                    'Suplemento (%)': supplement_pct,
                    'Tiempo Estándar (seg)': round(standard_time_seconds, 2),
                    'Tiempo Estándar (min)': round(standard_time_minutes, 4),
                    'Unidades/Hora': math.floor(units_per_hour),
                    'Unidades/Día': math.floor(units_per_day),
                })
                
            except KeyError as e:
                raise ExcelProcessingError(f"Missing required field in row {idx + 2}: {str(e)}")
            except Exception as e:
                raise ExcelProcessingError(f"Error processing row {idx + 2}: {str(e)}")
        
        return pd.DataFrame(results)
    
    def generate_excel_output(self, results_df: pd.DataFrame) -> io.BytesIO:
        """
        Generate an Excel file with formatted results
        """
        output = io.BytesIO()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Estimación de Producción"
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = results_df.columns.tolist()
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, row_data in enumerate(results_df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output)
        output.seek(0)
        
        return output
    
    def process_file(self, file_obj, filename: str) -> Tuple[io.BytesIO, Dict]:
        """
        Main processing method that orchestrates the entire workflow.
        
        Args:
            file_obj: File object to process
            filename: Name of the uploaded file
            
        Returns:
            Tuple of (Excel file as BytesIO, summary statistics dict)
        """
        # Validate file
        self.validate_file_extension(filename)
        
        # Read and validate Excel structure
        df = self.read_excel_file(file_obj)
        self.validate_dataframe_structure(df)
        
        # Calculate metrics
        results_df = self.calculate_metrics(df)
        
        # Generate output Excel
        output_file = self.generate_excel_output(results_df)
        
        # Generate summary statistics
        summary = {
            'total_operations': len(results_df),
            'avg_units_per_hour': round(results_df['Unidades/Hora'].mean(), 2),
            'avg_units_per_day': round(results_df['Unidades/Día'].mean(), 2),
        }
        
        return output_file, summary
