"""
Script to generate a sample Excel file for testing
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Sample data
data = {
    'Operador': ['Juan Pérez', 'María García', 'Carlos López', 'Ana Martínez', 'Pedro Rodríguez'],
    'Operación': ['Cortar tela', 'Coser manga izquierda', 'Coser manga derecha', 'Pegar botones', 'Planchar'],
    'Máquina': ['Cortadora industrial', 'Overlock', 'Overlock', 'Botonera', 'Plancha industrial'],
    'Tiempo 1': [45.2, 62.5, 61.8, 35.4, 28.6],
    'Tiempo 2': [46.1, 63.2, 62.1, 36.1, 29.1],
    'Tiempo 3': [44.8, 61.9, 63.0, 35.8, 28.3],
    'Tiempo 4': [45.5, 62.8, 61.5, 35.2, 28.9],
    'Tiempo 5': [45.9, 63.5, 62.4, 36.5, 28.7],
    'Suplemento': [15, 20, 20, 10, 12]
}

df = pd.DataFrame(data)

# Create Excel file
wb = Workbook()
ws = wb.active
ws.title = "Estudio de Tiempos"

# Header style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")

# Write headers
for col_idx, header in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Write data
for row_idx, row_data in enumerate(df.values, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save('sample_time_study.xlsx')
print("Sample file created: sample_time_study.xlsx")
